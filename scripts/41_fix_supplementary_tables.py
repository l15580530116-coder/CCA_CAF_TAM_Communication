#!/usr/bin/env python
"""Step 41: Fix supplementary tables content, generate Excel and CSV."""
import csv, shutil, os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path(r"e:\CCA")
TBL = BASE / "tables"
DST_CSV = BASE / "submission_CSBJ_UPLOAD_READY_FINAL/Supplementary_Tables_FINAL_CSV"
DST_CSV.mkdir(parents=True, exist_ok=True)
DST_XLSX = BASE / "submission_CSBJ_UPLOAD_READY_FINAL/Supplementary_Tables_CSBJ.xlsx"

# ============================================================
# 1. BUILD CORRECTED CSVs
# ============================================================

table_info = {}
all_rows = {}
all_headers = {}

def copy_csv(src, dst_name):
    """Copy CSV and record metadata."""
    shutil.copy2(src, DST_CSV / dst_name)
    with open(src, newline="", encoding="utf-8-sig") as f:
        reader = list(csv.reader(f))
    nrows = len(reader) - 1  # minus header
    ncols = len(reader[0]) if reader else 0
    all_headers[dst_name] = reader[0] if reader else []
    all_rows[dst_name] = reader[1:] if len(reader)>1 else []
    return nrows, ncols

def write_csv(dst_name, headers, rows):
    """Write a new CSV."""
    path = DST_CSV / dst_name
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)
    all_headers[dst_name] = headers
    all_rows[dst_name] = rows
    return len(rows), len(headers)

print("=== Building corrected supplementary tables ===\n")

# S1: Dataset summary
n1, c1 = copy_csv(TBL/"final_summary/final_number_verification_v3.csv",
                   "TableS1_dataset_summary.csv")
table_info["S1"] = ("Dataset summary", n1, c1,
                     "tables/final_summary/final_number_verification_v3.csv", "")
print(f"S1: {n1} rows, {c1} cols")

# S2: TCGA full DEG (9,291 significant)
n2, c2 = copy_csv(TBL/"DEG_TCGA_CHOL_significant.csv",
                   "TableS2_TCGA_CHOL_full_DEG_results.csv")
table_info["S2"] = ("TCGA-CHOL full DEG results", n2, c2,
                     "tables/DEG_TCGA_CHOL_significant.csv",
                     f"Should be 9,291 rows (significant DEGs)")
print(f"S2: {n2} rows, {c2} cols {'OK' if n2==9291 else 'CHECK: expected 9291'}")

# S3: GSE full DEG (7,735 significant)
n3, c3 = copy_csv(TBL/"DEG_GSE107943_paired_significant.csv",
                   "TableS3_GSE107943_full_DEG_results.csv")
table_info["S3"] = ("GSE107943 full DEG results", n3, c3,
                     "tables/DEG_GSE107943_paired_significant.csv",
                     f"Should be 7,735 rows (significant DEGs)")
print(f"S3: {n3} rows, {c3} cols {'OK' if n3==7735 else 'CHECK: expected 7735'}")

# S4: Validated IM/CAF/TAM
n4, c4 = copy_csv(TBL/"CCA_IM_CAF_TAM_DEGs_validated_same_direction.csv",
                   "TableS4_validated_IM_CAF_TAM_DEGs.csv")
table_info["S4"] = ("Validated IM/CAF/TAM DEGs", n4, c4,
                     "tables/CCA_IM_CAF_TAM_DEGs_validated_same_direction.csv", "")
print(f"S4: {n4} rows, {c4} cols")

# S5: Merged enrichment (GO up + GO down + KEGG)
enrich_rows = []
# GO BP up
with open(TBL/"enrichment/GO_BP_up_enrichment.csv", newline="", encoding="utf-8-sig") as f:
    for r in list(csv.DictReader(f)):
        r["direction"] = "up"; r["database"] = "GO_BP"
        enrich_rows.append(r)
# GO BP down
with open(TBL/"enrichment/GO_BP_down_enrichment.csv", newline="", encoding="utf-8-sig") as f:
    for r in list(csv.DictReader(f)):
        r["direction"] = "down"; r["database"] = "GO_BP"
        enrich_rows.append(r)
# KEGG up
kegg_path = TBL/"enrichment/KEGG_up_enrichment.csv"
if kegg_path.exists():
    with open(kegg_path, newline="", encoding="utf-8-sig") as f:
        for r in list(csv.DictReader(f)):
            r["direction"] = "up"; r["database"] = "KEGG"
            enrich_rows.append(r)

# Determine all field names
all_fields = []
for r in enrich_rows:
    for k in r:
        if k not in all_fields: all_fields.append(k)
# Ensure direction + database first
ordered = ["direction","database"] + [k for k in all_fields if k not in ("direction","database")]
enrich_csv_rows = [[r.get(k,"") for k in ordered] for r in enrich_rows]
n5, c5 = write_csv("TableS5_GO_KEGG_enrichment_results.csv", ordered, enrich_csv_rows)
table_info["S5"] = ("GO BP and KEGG enrichment results", n5, c5,
                     "tables/enrichment/ (merged GO_BP_up + GO_BP_down + KEGG_up)",
                     f"{n5} terms; KEGG down not available")
print(f"S5: {n5} rows, {c5} cols")

# S6: Merged ssGSEA (TCGA + GSE)
# TCGA tumor scores
tcga_rows = []
with open(TBL/"gsva/TCGA_CHOL_tumor_scores_with_survival.csv", newline="", encoding="utf-8-sig") as f:
    for r in list(csv.DictReader(f)):
        r["cohort"] = "TCGA-CHOL"; tcga_rows.append(r)
# GSE tumor scores
gse_rows = []
with open(TBL/"gsva/GSE107943_tumor_scores_with_survival.csv", newline="", encoding="utf-8-sig") as f:
    for r in list(csv.DictReader(f)):
        r["cohort"] = "GSE107943"; gse_rows.append(r)

all_s6 = tcga_rows + gse_rows
# Get all fields
s6_fields = ["cohort"]
for r in all_s6:
    for k in r:
        if k not in s6_fields: s6_fields.append(k)
s6_csv_rows = [[r.get(k,"") for k in s6_fields] for r in all_s6]
n6, c6 = write_csv("TableS6_ssGSEA_scores_TCGA_GSE107943.csv", s6_fields, s6_csv_rows)
table_info["S6"] = ("ssGSEA pathway scores per sample", n6, c6,
                     "tables/gsva/ (TCGA + GSE merged)",
                     f"TCGA={len(tcga_rows)} + GSE={len(gse_rows)} samples")
print(f"S6: {n6} rows, {c6} cols (TCGA={len(tcga_rows)}, GSE={len(gse_rows)})")

# S7: Immune correlations
n7, c7 = copy_csv(TBL/"immune/TCGA_aggr_vs_immune_correlation.csv",
                   "TableS7_immune_correlations.csv")
table_info["S7"] = ("Immune cell score correlations", n7, c7,
                     "tables/immune/TCGA_aggr_vs_immune_correlation.csv", "")
print(f"S7: {n7} rows, {c7} cols")

# S8: Checkpoint correlations
n8, c8 = copy_csv(TBL/"immune/TCGA_checkpoint_expression_correlation.csv",
                   "TableS8_checkpoint_correlations.csv")
table_info["S8"] = ("Checkpoint gene correlations", n8, c8,
                     "tables/immune/TCGA_checkpoint_expression_correlation.csv", "")
print(f"S8: {n8} rows, {c8} cols")

# S9: Single-cell scores
n9, c9 = copy_csv(TBL/"single_cell/GSE138709_score_by_celltype.csv",
                   "TableS9_single_cell_annotation_and_scores.csv")
table_info["S9"] = ("Single-cell scores by cell type", n9, c9,
                     "tables/single_cell/GSE138709_score_by_celltype.csv",
                     "Cell-type score summary")
print(f"S9: {n9} rows, {c9} cols")

# S10: CellChat LR pairs
n10, c10 = copy_csv(TBL/"cellchat/GSE138709_CAF_TAM_Epithelial_LR_pairs.csv",
                    "TableS10_CellChat_LR_pairs.csv")
table_info["S10"] = ("CellChat LR pairs", n10, c10,
                      "tables/cellchat/GSE138709_CAF_TAM_Epithelial_LR_pairs.csv", "")
print(f"S10: {n10} rows, {c10} cols")

# S11: Full 37-gene integrated scores
n11, c11 = copy_csv(TBL/"integrated/integrated_candidate_genes_all.csv",
                    "TableS11_integrated_hub_gene_evidence_scores_37genes.csv")
table_info["S11"] = ("Integrated hub gene evidence scores", n11, c11,
                      "tables/integrated/integrated_candidate_genes_all.csv",
                      f"Should be 37 genes")
print(f"S11: {n11} rows, {c11} cols {'OK' if n11==37 else 'CHECK: expected 37'}")

# S12: Merged Cox (TCGA + GSE)
cox_rows = []
with open(TBL/"clinical_relevance/TCGA_axis_score_survival_cox.csv", newline="", encoding="utf-8-sig") as f:
    for r in list(csv.DictReader(f)):
        r["cohort"] = "TCGA-CHOL"; cox_rows.append(r)
with open(TBL/"clinical_relevance/GSE107943_axis_score_survival_cox.csv", newline="", encoding="utf-8-sig") as f:
    for r in list(csv.DictReader(f)):
        r["cohort"] = "GSE107943"; cox_rows.append(r)
cox_fields = ["cohort"]
for r in cox_rows:
    for k in r:
        if k not in cox_fields: cox_fields.append(k)
cox_csv_rows = [[r.get(k,"") for k in cox_fields] for r in cox_rows]
n12, c12 = write_csv("TableS12_axis_score_Cox_TCGA_GSE107943.csv", cox_fields, cox_csv_rows)
table_info["S12"] = ("Axis score Cox regression", n12, c12,
                      "tables/clinical_relevance/ (TCGA + GSE merged)",
                      f"TCGA={len(cox_rows)-len([r for r in cox_rows if r['cohort']=='GSE107943'])} + GSE={len([r for r in cox_rows if r['cohort']=='GSE107943'])}")
print(f"S12: {n12} rows, {c12} cols (includes both TCGA and GSE)")

# S13: GSE26566
n13, c13 = copy_csv(TBL/"GSE26566_validation/GSE26566_hub_gene_expression_summary.csv",
                    "TableS13_GSE26566_validation.csv")
table_info["S13"] = ("GSE26566 hub gene expression", n13, c13,
                      "tables/GSE26566_validation/GSE26566_hub_gene_expression_summary.csv", "")
print(f"S13: {n13} rows, {c13} cols")

# S14: Druggability
n14, c14 = copy_csv(TBL/"drug_screening/prioritized_therapeutic_targets.csv",
                    "TableS14_druggability_targets.csv")
table_info["S14"] = ("Druggability targets", n14, c14,
                      "tables/drug_screening/prioritized_therapeutic_targets.csv", "")
print(f"S14: {n14} rows, {c14} cols")

# S15: Candidate drugs
n15, c15 = copy_csv(TBL/"drug_screening/candidate_drugs_prioritized.csv",
                    "TableS15_candidate_drug_table.csv")
table_info["S15"] = ("Candidate drug table", n15, c15,
                      "tables/drug_screening/candidate_drugs_prioritized.csv", "")
print(f"S15: {n15} rows, {c15} cols")

# S16: Docking parameters
n16, c16 = copy_csv(TBL/"docking/docking_results_all_pairs_merged.csv",
                    "TableS16_docking_parameters.csv")
table_info["S16"] = ("Docking parameters", n16, c16,
                      "tables/docking/docking_results_all_pairs_merged.csv", "")
print(f"S16: {n16} rows, {c16} cols")

# S17: Docking neighbor residues
n17, c17 = copy_csv(TBL/"docking/docking_neighbor_residues_4A.csv",
                    "TableS17_docking_neighbor_residues.csv")
table_info["S17"] = ("Docking neighbor residues", n17, c17,
                      "tables/docking/docking_neighbor_residues_4A.csv", "")
print(f"S17: {n17} rows, {c17} cols")

# ============================================================
# 2. BUILD EXCEL
# ============================================================
print("\n=== Building Excel ===")

wb = Workbook()
# Remove default sheet
wb.remove(wb.active)

header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
cell_font = Font(name="Arial", size=9)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"))

sheet_names = {
    "S1": "S1_Dataset", "S2": "S2_TCGA_DEG", "S3": "S3_GSE_DEG",
    "S4": "S4_Validated_DEGs", "S5": "S5_Enrichment", "S6": "S6_ssGSEA",
    "S7": "S7_ImmuneCorr", "S8": "S8_CheckpointCorr", "S9": "S9_scRNA",
    "S10": "S10_CellChatLR", "S11": "S11_HubScores", "S12": "S12_Cox",
    "S13": "S13_GSE26566", "S14": "S14_Druggability", "S15": "S15_Drugs",
    "S16": "S16_DockingParams", "S17": "S17_DockingResidues",
}

# Index sheet
ws_idx = wb.create_sheet("Index", 0)
idx_headers = ["Table ID", "Sheet name", "Description", "Rows", "Columns",
               "Source file", "Notes"]
for c, h in enumerate(idx_headers, 1):
    cell = ws_idx.cell(row=1, column=c, value=h)
    cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border
for r, (s_id, (desc, nr, nc, src, note)) in enumerate(table_info.items(), 2):
    vals = [f"Table {s_id}", sheet_names[s_id], desc, nr, nc, src, note]
    for c, v in enumerate(vals, 1):
        cell = ws_idx.cell(row=r, column=c, value=v)
        cell.font = cell_font; cell.border = thin_border
# Column widths
for c, w in enumerate([12, 18, 50, 8, 8, 55, 40], 1):
    ws_idx.column_dimensions[get_column_letter(c)].width = w
ws_idx.freeze_panes = "A2"
ws_idx.auto_filter.ref = f"A1:G{len(table_info)+1}"

# Data sheets
for s_id in [f"S{i}" for i in range(1,18)]:
    sn = sheet_names[s_id]
    ws = wb.create_sheet(sn)
    fname = [f for f in os.listdir(DST_CSV) if f.startswith(f"Table{s_id}_")][0]
    headers = all_headers[fname]
    rows_data = all_rows[fname]

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    for r, row in enumerate(rows_data, 2):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = cell_font; cell.border = thin_border
    ws.freeze_panes = "A2"
    if len(rows_data) > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows_data)+1}"
    # Auto-width (approximate)
    for c in range(1, min(len(headers)+1, 20)):
        ws.column_dimensions[get_column_letter(c)].width = 14

wb.save(str(DST_XLSX))
print(f"Saved: {DST_XLSX.name}")

# ============================================================
# 3. QC SUMMARY
# ============================================================
print(f"\n=== QC Summary ===")
print(f"S2 TCGA DEG: {n2} rows (expected 9291) — {'OK' if n2==9291 else 'CHECK'}")
print(f"S3 GSE DEG: {n3} rows (expected 7735) — {'OK' if n3==7735 else 'CHECK'}")
print(f"S5 Enrichment: {n5} terms (from GO up+down+KEGG up)")
print(f"S6 ssGSEA: TCGA={len(tcga_rows)} + GSE={len(gse_rows)} samples")
print(f"S11 Hub genes: {n11} genes (expected 37)")
print(f"S12 Cox: {n12} rows (TCGA+GSE combined)")
print(f"S16 Docking: {n16} pairs")
print(f"S17 Residues: {n17} residues (total across all pairs)")
print(f"Excel: {DST_XLSX.name}")
print(f"CSV dir: {DST_CSV}")
